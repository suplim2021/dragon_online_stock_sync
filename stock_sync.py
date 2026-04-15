"""
stock_sync.py — Dragon Fastener Online Stock Sync
บริษัท ศรีวัฒนะบุญ ซัพพลายส์ จำกัด

Usage (auto-detect):
    python3 stock_sync.py

Usage (manual override):
    python3 stock_sync.py <shopee_file> <lazada_file> [tiktok_file]

Auto-detect picks the newest file matching each prefix in the same folder:
    Shopee  → mass_update_sales_info_*.xlsx
    Lazada  → pricestock*.xlsx
    TikTok  → Tiktoksellercenter_*.xlsx  (optional — skipped if not found)

Fixed:
    Warehouse stock: บริษัท ศรีวัฒนะบุญ ซัพพลายส์ จำกัด_SimpleInventoryReport_.xlsx

Output folders:
    Reports/   ← StockSync report with all tabs
    Uploads/   ← Shopee_updated + Lazada_updated [+ TikTok_updated]
"""

import sys, zipfile, re, io, math, os, glob
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─── INPUT FILES ──────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
STOCK_FILE  = os.path.join(BASE_DIR, "บริษัท ศรีวัฒนะบุญ ซัพพลายส์ จำกัด_SimpleInventoryReport_.xlsx")

def find_latest(pattern):
    """Return the most recently modified file matching glob pattern, or None."""
    matches = glob.glob(os.path.join(BASE_DIR, pattern))
    if not matches:
        return None
    return max(matches, key=os.path.getmtime)

if len(sys.argv) >= 3:
    # Manual override
    SHOPEE_FILE = os.path.join(BASE_DIR, sys.argv[1])
    LAZADA_FILE = os.path.join(BASE_DIR, sys.argv[2])
    TIKTOK_FILE = os.path.join(BASE_DIR, sys.argv[3]) if len(sys.argv) >= 4 else find_latest("Tiktoksellercenter_*.xlsx")
else:
    # Auto-detect newest file per platform
    SHOPEE_FILE = find_latest("mass_update_sales_info_*.xlsx")
    LAZADA_FILE = find_latest("pricestock*.xlsx")
    TIKTOK_FILE = find_latest("Tiktoksellercenter_*.xlsx")

# Validate required files
missing = []
if not SHOPEE_FILE or not os.path.exists(SHOPEE_FILE):
    missing.append("Shopee  (mass_update_sales_info_*.xlsx)")
if not LAZADA_FILE or not os.path.exists(LAZADA_FILE):
    missing.append("Lazada  (pricestock*.xlsx)")
if missing:
    print("❌ ไม่พบไฟล์:")
    for m in missing: print(f"   • {m}")
    print("วางไฟล์ไว้ในโฟลเดอร์เดียวกับ stock_sync.py แล้วลองใหม่")
    sys.exit(1)

# Print detected files
print("📂 ไฟล์ที่ใช้:")
print(f"   Shopee : {os.path.basename(SHOPEE_FILE)}")
print(f"   Lazada : {os.path.basename(LAZADA_FILE)}")
if TIKTOK_FILE and os.path.exists(TIKTOK_FILE):
    print(f"   TikTok : {os.path.basename(TIKTOK_FILE)}")
else:
    TIKTOK_FILE = None
    print(f"   TikTok : (ไม่พบ — ข้าม)")

NOW         = datetime.now().strftime("%Y%m%d_%H%M")
OUT_SYNC    = os.path.join(BASE_DIR, f"Reports/StockSync_{NOW}.xlsx")
OUT_SHOPEE  = os.path.join(BASE_DIR, f"Uploads/Shopee_updated_{NOW}.xlsx")
OUT_LAZADA  = os.path.join(BASE_DIR, f"Uploads/Lazada_updated_{NOW}.xlsx")
OUT_TIKTOK  = os.path.join(BASE_DIR, f"Uploads/TikTok_updated_{NOW}.xlsx")

os.makedirs(os.path.join(BASE_DIR, "Reports"), exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "Uploads"), exist_ok=True)

# ─── SKU TABLE ────────────────────────────────────────────────────────────────
# ppb = pcs per bag, ctn = pcs per carton (full crate)
CTN_TABLE = {
    'DG10-16X16SD':  {'ppb': 100, 'ctn': 4000},
    'DG10-16X20SD':  {'ppb': 100, 'ctn': 4000},
    'DG12-14X20SD':  {'ppb': 100, 'ctn': 2500},
    'DG12-14X20SD-N':{'ppb': 100, 'ctn': 2500},
    'DG12-14X30SD':  {'ppb': 100, 'ctn': 2000},
    'DG15-15X20SD':  {'ppb': 100, 'ctn': 2500},
    'DG12-14X48SD':  {'ppb': 100, 'ctn': 1200},
    'DG12-14X48SD-B':{'ppb': 100, 'ctn': 1200},
    'DG12-14X48SD-C2':{'ppb': 100, 'ctn': 1200},
    'DG12-14X55SD':  {'ppb': 100, 'ctn': 1000},
    'DG12-14X65SD':  {'ppb': 100, 'ctn': 1000},
    'DG12-14X75SD':  {'ppb': 100, 'ctn': 800},
    'DG12-14X75SD-B':{'ppb': 100, 'ctn': 800},
    'DG12-14X85SD':  {'ppb': 100, 'ctn': 800},
    'DG12-14X110SD': {'ppb': 50,  'ctn': 600},
    'DG12-14X130SD': {'ppb': 50,  'ctn': 600},
    'DG10-12X20T':   {'ppb': 100, 'ctn': 4000},
    'DG12-11X50T':   {'ppb': 100, 'ctn': 1200},
    'DG12-11X50T-B': {'ppb': 100, 'ctn': 1200},
    'DG12-11X65T':   {'ppb': 100, 'ctn': 1000},
    'DG12-11X75T':   {'ppb': 100, 'ctn': 800},
    'DG10-24X16WF':  {'ppb': 100, 'ctn': 5000},
    'DG10-24X22WF':  {'ppb': 100, 'ctn': 5000},
    # Bag-only products (no crate)
    'DG-#8':              {'ppb': 1,   'ctn': None},   # หัวบล็อก — ขายตัวเดี่ยว
    'DG-BONDED-WASHER':   {'ppb': 100, 'ctn': None},   # แหวนอะลูมิเนียม 16mm
    'DG-BONDED-WASHER-25':{'ppb': 100, 'ctn': None},   # แหวนอะลูมิเนียม 25mm
    'DG-EPDM-DOME':       {'ppb': 100, 'ctn': None},   # แหวนยางโดม
    'DG-LOUVER':          {'ppb': 20,  'ctn': None},   # ลูฟเวอร์
    # S- series — ขายเป็นหลอด (ppb=1)
    'S-DBOND-BLACK':      {'ppb': 1,   'ctn': None},
    'S-DBOND-CLEAR':      {'ppb': 1,   'ctn': None},
    'S-DBOND-LIGHTGREY':  {'ppb': 1,   'ctn': None},
    'S-DBOND-WHITE':      {'ppb': 1,   'ctn': None},
    'S-DBOND-ZINC':       {'ppb': 1,   'ctn': None},
}

def get_base_info(sku):
    """Return (base_sku, is_base, ppb, ctn).
    is_base=True  → plain SKU (no color suffix)
    is_base=False → colored variant (has suffix like -B, -C2, -NaturalBrown)
    is_base=None  → unknown SKU
    """
    if sku in CTN_TABLE:
        return sku, True, CTN_TABLE[sku]['ppb'], CTN_TABLE[sku]['ctn']
    for base, info in CTN_TABLE.items():
        if sku.startswith(base + '-'):
            return base, False, info['ppb'], info['ctn']
    ppb = 50 if ('X110' in sku.upper() or 'X130' in sku.upper()) else 100
    return sku, None, ppb, None

def calc_suggest(qty, unit, sku):
    """Calculate suggested stock for Shopee, Lazada, and TikTok.
    Returns (suggest_shopee, suggest_lazada, suggest_tiktok).

    TikTok rules:
      Bag  — regular: same as Lazada | colored: floor(lazada × 0.6)
      Crate — regular: same as Lazada | colored: max(0, suggest_ctn - 1)
    """
    base, is_base, ppb, ctn = get_base_info(sku)
    colored = not is_base   # True if colored variant (suffix exists)

    if unit == 'ลัง':
        if ctn is None:
            return 0, 0, 0
        full_crates = math.floor(qty / ctn)
        remainder   = qty % ctn
        # Safety: if exactly 1 full crate and no remainder → suggest 0 (don't sell last crate)
        suggest_ctn = 0 if (full_crates == 1 and remainder == 0) else full_crates
        lz = 0 if colored else math.floor(suggest_ctn * 0.6)
        # TikTok: colored → -1 crate (min 0); regular → same as Lazada
        tk = max(0, suggest_ctn - 1) if colored else lz
        return suggest_ctn, lz, tk
    else:
        # Bag listing
        safe = max(0, math.floor(qty / ppb) - 1)   # keep 1 bag as safety buffer
        sh   = min(safe, 50)                         # Shopee cap = 50
        lz   = 0 if (colored and safe < 5) else math.floor(sh * 0.6)
        # TikTok: colored → floor(lz × 0.6) (another tier down); regular → same as Lazada
        tk   = math.floor(lz * 0.6) if colored else lz
        return sh, lz, tk

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def to_int(v):
    try: return int(v)
    except: return None

def detect_unit(name):
    return 'ลัง' if 'ลัง' in str(name) else 'ถุง'

def load_shopee(path):
    """Load Shopee Excel, patching the activePane XML bug."""
    with open(path, 'rb') as f:
        raw = f.read()
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(raw)) as zin, \
         zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith('xl/worksheets/'):
                xml  = data.decode('utf-8')
                xml  = re.sub(r'\bactivePane="[^"]*"', '', xml)
                data = xml.encode('utf-8')
            zout.writestr(item, data)
    buf.seek(0)
    return openpyxl.load_workbook(buf)

# ─── COLOUR PALETTE ───────────────────────────────────────────────────────────
C_ORANGE  = PatternFill("solid", fgColor="FF6600")
C_YELLOW  = PatternFill("solid", fgColor="FFD700")
C_BLUE    = PatternFill("solid", fgColor="4472C4")
C_LBLUE   = PatternFill("solid", fgColor="BDD7EE")
C_LBLUE2  = PatternFill("solid", fgColor="DDEEFF")
C_WHITE   = PatternFill("solid", fgColor="FFFFFF")
C_LGREEN  = PatternFill("solid", fgColor="E2EFDA")
C_LRED    = PatternFill("solid", fgColor="FCE4D6")
C_LYELLOW = PatternFill("solid", fgColor="FFFF99")
C_LGRAY   = PatternFill("solid", fgColor="F2F2F2")
C_TIKTOK  = PatternFill("solid", fgColor="2D2D2D")  # TikTok dark
CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=True)

def flag_fill(f):
    if '✅' in f:  return C_LGREEN
    if '⚠️' in f: return C_LYELLOW
    if '🔴' in f:  return C_LRED
    if '⏭️' in f: return C_LBLUE2
    return C_LGRAY

# ─── READ WAREHOUSE STOCK ─────────────────────────────────────────────────────
print("📦 Reading warehouse stock…")
wb_s  = openpyxl.load_workbook(STOCK_FILE, data_only=True)
ws_s  = wb_s['Inventory Warehouse']
stock_rows = []
for row in ws_s.iter_rows(min_row=8, values_only=True):
    if row[0] is None or not isinstance(row[0], (int, float)):
        continue
    stock_rows.append(row)

stock_exact = {str(r[2]).strip(): {'name': r[1], 'qty': float(r[4] or 0)} for r in stock_rows if r[2]}
stock_lower = {k.lower(): v for k, v in stock_exact.items()}
print(f"   {len(stock_rows)} rows, {len(stock_exact)} unique SKUs")

def lookup(sku):
    if sku in stock_exact: return stock_exact[sku]
    if sku.lower() in stock_lower: return stock_lower[sku.lower()]
    return None

# ─── READ SHOPEE ──────────────────────────────────────────────────────────────
print("🟠 Reading Shopee…")
wb_sh = load_shopee(SHOPEE_FILE)
shopee_rows = []
for i, r in enumerate(wb_sh.active.iter_rows(min_row=7, values_only=True), 7):
    sku = str(r[5]).strip() if r[5] else ''   # Col F = เลข SKU (variation)
    if not sku:
        sku = str(r[4]).strip() if r[4] else ''  # Col E = Parent SKU (fallback)
    if sku:
        shopee_rows.append((i, sku, str(r[1]).strip(), to_int(r[8])))
print(f"   {len(shopee_rows)} rows")

# ─── READ LAZADA ──────────────────────────────────────────────────────────────
print("🟡 Reading Lazada…")
wb_lz = openpyxl.load_workbook(LAZADA_FILE)
lazada_rows = [
    (i, str(r[12]).strip(), str(r[2]).strip(), to_int(r[7]))
    for i, r in enumerate(wb_lz.active.iter_rows(min_row=5, values_only=True), 5)
    if r[12]
]
print(f"   {len(lazada_rows)} rows")

# ─── READ TIKTOK (optional) ───────────────────────────────────────────────────
tiktok_rows = []
wb_tk = None
if TIKTOK_FILE:
    print("⬛ Reading TikTok…")
    wb_tk = openpyxl.load_workbook(TIKTOK_FILE)
    ws_tk = wb_tk['Template']
    for i, r in enumerate(ws_tk.iter_rows(min_row=6, values_only=True), 6):
        sku = str(r[7]).strip() if r[7] else ''   # Col H = seller_sku
        if sku:
            tiktok_rows.append((i, sku, str(r[2]).strip(), to_int(r[6])))
    print(f"   {len(tiktok_rows)} rows")

# ─── CALCULATE SHOPEE ─────────────────────────────────────────────────────────
# Rule: [SALE][ยกลัง] → update normally (carton stock)
#        [ยกลัง] only AND the same SKU also has a [SALE] version → skip (old listing)
#        [ยกลัง] only AND no [SALE] counterpart → update normally
print("\n🔢 Calculating Shopee…")

# Step 1: find which SKUs have an ACTIVE [SALE][ยกลัง] listing (stock > 0)
sale_skus_active = set()
for (ri, sku, name, curr) in shopee_rows:
    if '[SALE]' in name and 'ลัง' in name and curr and curr > 0:
        sale_skus_active.add(sku)

# Step 2: process each row
shopee_data = []
for (ri, sku, name, curr) in shopee_rows:
    unit = detect_unit(name)

    # [SALE][ยกลัง] แต่สต็อกเดิม = 0 → listing ไม่ active → ข้าม
    if '[SALE]' in name and 'ลัง' in name and (not curr or curr == 0):
        info = lookup(sku)
        shopee_data.append({
            'ri': ri, 'sku': sku, 'name': name, 'unit': unit,
            'qty': info['qty'] if info else None,
            'curr': curr, 'suggest_sh': 0, 'suggest_lz': 0, 'suggest_tk': 0,
            'flag': '⏭️ ข้าม (SALE ไม่ active)'
        })
        continue

    # [ยกลัง] ปกติ → skip เฉพาะถ้า SKU นี้มี [SALE][ยกลัง] ที่ active อยู่
    if 'ลัง' in name and '[SALE]' not in name and sku in sale_skus_active:
        info = lookup(sku)
        shopee_data.append({
            'ri': ri, 'sku': sku, 'name': name, 'unit': unit,
            'qty': info['qty'] if info else None,
            'curr': curr, 'suggest_sh': curr or 0, 'suggest_lz': 0, 'suggest_tk': 0,
            'flag': '⏭️ ข้าม'
        })
        continue

    # All other rows: calculate normally
    info = lookup(sku)
    if info is None:
        flag = '❓ สินค้าอื่น' if (not sku.upper().startswith('DG') and not sku.upper().startswith('S-')) else '❓ ไม่พบใน Stock'
        shopee_data.append({'ri': ri, 'sku': sku, 'name': name, 'unit': unit,
                            'qty': None, 'curr': curr, 'suggest_sh': 0, 'suggest_lz': 0, 'suggest_tk': 0, 'flag': flag})
        continue

    qty = info['qty']
    sh, lz, tk = calc_suggest(qty, unit, sku)
    flag = '✅ โอเค' if sh == curr else ('⚠️ ต้องปรับ' if sh > 0 else '🔴 หยุดขาย')
    shopee_data.append({'ri': ri, 'sku': sku, 'name': name, 'unit': unit,
                        'qty': qty, 'curr': curr, 'suggest_sh': sh, 'suggest_lz': lz, 'suggest_tk': tk, 'flag': flag})

# ─── CALCULATE LAZADA ─────────────────────────────────────────────────────────
print("🔢 Calculating Lazada…")
lazada_data = []
for (ri, sku, name, curr) in lazada_rows:
    unit = detect_unit(name)
    info = lookup(sku)
    if info is None:
        flag = '❓ สินค้าอื่น' if (not sku.upper().startswith('DG') and not sku.upper().startswith('S-')) else '❓ ไม่พบใน Stock'
        lazada_data.append({'ri': ri, 'sku': sku, 'name': name, 'unit': unit,
                            'qty': None, 'curr': curr, 'suggest_sh': 0, 'suggest_lz': 0, 'suggest_tk': 0, 'flag': flag})
        continue
    qty = info['qty']
    sh, lz, tk = calc_suggest(qty, unit, sku)
    flag = '✅ โอเค' if lz == curr else ('⚠️ ต้องปรับ' if lz > 0 else '🔴 หยุดขาย')
    lazada_data.append({'ri': ri, 'sku': sku, 'name': name, 'unit': unit,
                        'qty': qty, 'curr': curr, 'suggest_sh': sh, 'suggest_lz': lz, 'suggest_tk': tk, 'flag': flag})

# ─── CALCULATE TIKTOK ─────────────────────────────────────────────────────────
tiktok_data = []
if tiktok_rows:
    print("🔢 Calculating TikTok…")
    for (ri, sku, name, curr) in tiktok_rows:
        unit = detect_unit(name)
        info = lookup(sku)
        if info is None:
            flag = '❓ สินค้าอื่น' if (not sku.upper().startswith('DG') and not sku.upper().startswith('S-')) else '❓ ไม่พบใน Stock'
            tiktok_data.append({'ri': ri, 'sku': sku, 'name': name, 'unit': unit,
                                'qty': None, 'curr': curr, 'suggest_sh': 0, 'suggest_lz': 0, 'suggest_tk': 0, 'flag': flag})
            continue
        qty = info['qty']
        sh, lz, tk = calc_suggest(qty, unit, sku)
        flag = '✅ โอเค' if tk == curr else ('⚠️ ต้องปรับ' if tk > 0 else '🔴 หยุดขาย')
        tiktok_data.append({'ri': ri, 'sku': sku, 'name': name, 'unit': unit,
                            'qty': qty, 'curr': curr, 'suggest_sh': sh, 'suggest_lz': lz, 'suggest_tk': tk, 'flag': flag})

# Summary
from collections import Counter
for label, data in [('Shopee', shopee_data), ('Lazada', lazada_data)] + ([('TikTok', tiktok_data)] if tiktok_data else []):
    c = Counter(d['flag'] for d in data)
    print(f"   {label}: {dict(c)}")

# ─── BUILD StockSync REPORT ───────────────────────────────────────────────────
print("\n📊 Building StockSync report…")

def write_sync_sheet(ws, data, platform, date_str):
    if platform == 'Shopee':
        sc, clr, fc, title = 'suggest_sh', C_ORANGE, "FFFFFF", "🟠 Shopee"
    elif platform == 'TikTok':
        sc, clr, fc, title = 'suggest_tk', C_TIKTOK, "FFFFFF", "⬛ TikTok"
    else:  # Lazada
        sc, clr, fc, title = 'suggest_lz', C_YELLOW, "000000", "🟡 Lazada"

    # Title row
    ws.merge_cells('A1:P1')
    t = ws.cell(row=1, column=1, value=f"{title} Stock Sync — {date_str}")
    t.fill = clr
    t.font = Font(bold=True, size=13, name="Calibri", color=fc)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # Header row
    hdrs = ['สถานะ', 'เลข SKU', 'ชื่อตัวเลือก', 'ชื่อสินค้า', 'หน่วย', 'ประเภท',
            'ราคาปัจจุบัน', 'สต็อกเดิม', 'สต็อกจริง(ตัว)', '🔗 สต็อก(ลิ้งค์)',
            'pcs/ถุง', 'แนะนำ(ถุง/ลัง)', 'ใส่ Shopee', 'ใส่ Lazada', 'ใส่ TikTok', 'ส่วนต่าง']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = clr
        c.font = Font(bold=True, size=9, name="Calibri", color=fc)
        c.alignment = CENTER
    ws.row_dimensions[2].height = 30

    # Data rows — sorted by SKU for readability (upload files are unaffected)
    for r_idx, d in enumerate(sorted(data, key=lambda x: x['sku']), start=3):
        fl  = flag_fill(d['flag'])
        _, is_base, ppb, _ = get_base_info(d['sku'])
        suggest = d[sc]
        curr    = d['curr']
        diff    = (suggest - curr) if (suggest is not None and curr is not None) else None
        type_lbl = 'ไม่มีสี' if is_base else ('สี' if is_base is False else 'อื่นๆ')

        vals = [d['flag'], d['sku'], '', d['name'], d['unit'], type_lbl,
                None, curr, d['qty'], None, ppb, suggest,
                d['suggest_sh'], d['suggest_lz'], d.get('suggest_tk', 0), diff]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r_idx, column=ci, value=v)
            c.fill = fl
            c.font = Font(size=9, name="Calibri")
            c.alignment = LEFT if ci in (2, 3, 4) else CENTER

        # Column J: IFERROR INDEX/MATCH formula linking to Stock sheet
        fm = ws.cell(row=r_idx, column=10,
                     value=f"=IFERROR(INDEX('📦 Stock'!$D:$D,MATCH(B{r_idx},'📦 Stock'!$C:$C,0)),\"❓\")")
        fm.fill = C_LBLUE
        fm.font = Font(size=9, name="Calibri", color="1F4E79")
        fm.alignment = CENTER

    # Column widths
    widths = [14, 22, 18, 40, 8, 10, 10, 10, 12, 12, 8, 12, 10, 10, 10, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A3'


date_label = datetime.now().strftime("%d/%m/%Y")
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

write_sync_sheet(wb_out.create_sheet('🟠 ShopeeSync'), shopee_data, 'Shopee', date_label)
write_sync_sheet(wb_out.create_sheet('🟡 LazadaSync'), lazada_data, 'Lazada', date_label)
if tiktok_data:
    write_sync_sheet(wb_out.create_sheet('⬛ TikTokSync'), tiktok_data, 'TikTok', date_label)

# ─── 📦 Stock sheet ───────────────────────────────────────────────────────────
ws_stk = wb_out.create_sheet('📦 Stock')
ws_stk.merge_cells('A1:G1')
t = ws_stk.cell(row=1, column=1, value=f"📦 Stock Snapshot — {date_label}")
t.fill = C_BLUE
t.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
t.alignment = CENTER

for ci, h in enumerate(['ลำดับ', 'ชื่อสินค้า', 'รหัสสินค้า', 'จำนวนคงคลัง', 'หน่วย', 'ราคาเฉลี่ย', 'มูลค่า'], 1):
    c = ws_stk.cell(row=2, column=ci, value=h)
    c.fill = C_BLUE
    c.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
    c.alignment = CENTER

for ri, row in enumerate(stock_rows, start=3):
    # Reorder: original [3]=ราคา [4]=จำนวน → put จำนวน(col D) before ราคา
    reordered = (row[0], row[1], row[2], row[4], row[5], row[3], row[6])
    for ci, v in enumerate(reordered, 1):
        c = ws_stk.cell(row=ri, column=ci, value=v)
        c.fill = C_LGRAY if ri % 2 == 0 else C_WHITE
        c.font = Font(size=9, name="Calibri")
        c.alignment = LEFT if ci == 2 else CENTER

ws_stk.column_dimensions['B'].width = 40
ws_stk.column_dimensions['C'].width = 25
ws_stk.column_dimensions['D'].width = 14
ws_stk.freeze_panes = 'A3'

# ─── Raw tabs ─────────────────────────────────────────────────────────────────
for shname, wb_src in [('🛒 Shopee', wb_sh), ('🟡 Lazada', wb_lz)]:
    ws_raw = wb_out.create_sheet(shname)
    for row in wb_src.active.iter_rows(values_only=True):
        ws_raw.append(list(row))

if wb_tk:
    ws_raw_tk = wb_out.create_sheet('⬛ TikTok')
    ws_tk2 = wb_tk['Template']
    for row in ws_tk2.iter_rows(values_only=True):
        ws_raw_tk.append(list(row))

wb_out.save(OUT_SYNC)
print(f"✅ StockSync saved: Reports/StockSync_{NOW}.xlsx")

# ─── UPDATE SHOPEE FILE ───────────────────────────────────────────────────────
print("\n🟠 Updating Shopee file…")
wb_sh2   = load_shopee(SHOPEE_FILE)
skip_sh  = upd_sh = zero_sh = 0
for d in shopee_data:
    if '⏭️' in d['flag'] or '❓ สินค้าอื่น' in d['flag']:
        skip_sh += 1
        continue
    if '❓ ไม่พบใน Stock' in d['flag']:
        wb_sh2.active.cell(row=d['ri'], column=9, value=0)
        zero_sh += 1
        continue
    wb_sh2.active.cell(row=d['ri'], column=9, value=d['suggest_sh'])
    upd_sh += 1
wb_sh2.save(OUT_SHOPEE)
print(f"   Updated {upd_sh} rows | ใส่ 0: {zero_sh} rows | Skipped {skip_sh} rows")
print(f"   Saved: Uploads/Shopee_updated_{NOW}.xlsx")

# ─── UPDATE LAZADA FILE ───────────────────────────────────────────────────────
print("\n🟡 Updating Lazada file…")
wb_lz2   = openpyxl.load_workbook(LAZADA_FILE)
skip_lz  = upd_lz = zero_lz = 0
for d in lazada_data:
    if '❓ สินค้าอื่น' in d['flag']:
        skip_lz += 1
        continue
    if '❓ ไม่พบใน Stock' in d['flag']:
        wb_lz2.active.cell(row=d['ri'], column=8, value=0)
        zero_lz += 1
        continue
    wb_lz2.active.cell(row=d['ri'], column=8, value=d['suggest_lz'])
    upd_lz += 1
wb_lz2.save(OUT_LAZADA)
print(f"   Updated {upd_lz} rows | ใส่ 0: {zero_lz} rows | Skipped {skip_lz} rows")
print(f"   Saved: Uploads/Lazada_updated_{NOW}.xlsx")

# ─── UPDATE TIKTOK FILE ───────────────────────────────────────────────────────
if tiktok_data and TIKTOK_FILE:
    print("\n⬛ Updating TikTok file…")
    wb_tk3   = openpyxl.load_workbook(TIKTOK_FILE)
    ws_tk3   = wb_tk3['Template']
    skip_tk  = upd_tk = zero_tk = 0
    for d in tiktok_data:
        if '❓ สินค้าอื่น' in d['flag']:
            skip_tk += 1
            continue
        if '❓ ไม่พบใน Stock' in d['flag']:
            ws_tk3.cell(row=d['ri'], column=7, value=0)  # Col G = quantity
            zero_tk += 1
            continue
        ws_tk3.cell(row=d['ri'], column=7, value=d['suggest_tk'])
        upd_tk += 1
    wb_tk3.save(OUT_TIKTOK)
    print(f"   Updated {upd_tk} rows | ใส่ 0: {zero_tk} rows | Skipped {skip_tk} rows")
    print(f"   Saved: Uploads/TikTok_updated_{NOW}.xlsx")

# ─── QUICK VERIFY ─────────────────────────────────────────────────────────────
print("\n🔍 Verify Stock sheet columns:")
wb_v  = openpyxl.load_workbook(OUT_SYNC, data_only=True)
ws_v  = wb_v['📦 Stock']
hdrs  = [ws_v.cell(row=2, column=c).value for c in range(1, 8)]
print(f"   D = {hdrs[3]}   E = {hdrs[4]}")
for row in ws_v.iter_rows(min_row=3, max_row=5, values_only=True):
    print(f"   SKU={str(row[2]):25s}  D(จำนวน)={row[3]}  E(หน่วย)={row[4]}")

print("\n✅ Done!")
